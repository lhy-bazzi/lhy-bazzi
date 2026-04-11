"""Excel / CSV parser — XLSX, XLS, CSV → structured table elements."""

from __future__ import annotations

import asyncio
import os
from typing import Any

import chardet
from loguru import logger

from app.models.document import DocumentMetadata, ParsedDocument, DocumentElement
from app.models.enums import ElementType, FileType
from app.services.parsing.base import BaseParser
from app.utils.exceptions import ParseError

# Max rows per element to avoid huge single chunks
_MAX_ROWS_PER_ELEMENT = 50


class ExcelParser(BaseParser):
    """Parse XLSX / XLS / CSV files into TABLE elements."""

    def supported_types(self) -> list[FileType]:
        return [FileType.XLSX, FileType.XLS, FileType.CSV]

    async def parse(self, file_path: str, **kwargs: Any) -> ParsedDocument:
        doc_id = kwargs.get("doc_id", self._gen_id())
        filename = os.path.basename(file_path)
        ext = os.path.splitext(filename)[1].lower()
        logger.info("Parsing spreadsheet '{}' (ext={})", filename, ext)

        try:
            if ext == ".csv":
                elements = await asyncio.to_thread(self._parse_csv, file_path)
            elif ext == ".xls":
                elements = await asyncio.to_thread(self._parse_xls, file_path)
            else:
                elements = await asyncio.to_thread(self._parse_xlsx, file_path)
        except Exception as exc:
            raise ParseError(f"Spreadsheet parse failed: {exc}") from exc

        metadata = DocumentMetadata(file_size_bytes=self._file_size(file_path))
        return self._build_document(
            doc_id=doc_id,
            filename=filename,
            file_type=ext.lstrip("."),
            elements=elements,
            parse_engine="openpyxl" if ext != ".xls" else "xlrd",
            metadata=metadata,
        )

    # ---- XLSX ----

    def _parse_xlsx(self, file_path: str) -> list[DocumentElement]:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        elements: list[DocumentElement] = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            elements.extend(self._rows_to_elements(rows, sheet_name=sheet))
        wb.close()
        return elements

    # ---- XLS ----

    def _parse_xls(self, file_path: str) -> list[DocumentElement]:
        import xlrd

        wb = xlrd.open_workbook(file_path)
        elements: list[DocumentElement] = []
        for sheet in wb.sheets():
            rows = []
            for r in range(sheet.nrows):
                rows.append(tuple(sheet.cell_value(r, c) for c in range(sheet.ncols)))
            elements.extend(self._rows_to_elements(rows, sheet_name=sheet.name))
        return elements

    # ---- CSV ----

    def _parse_csv(self, file_path: str) -> list[DocumentElement]:
        import pandas as pd

        # Detect encoding
        with open(file_path, "rb") as f:
            raw = f.read(8192)
        detected = chardet.detect(raw)
        encoding = detected.get("encoding") or "utf-8"

        df = pd.read_csv(file_path, encoding=encoding, on_bad_lines="skip")
        rows = [tuple(df.columns)] + [tuple(row) for row in df.itertuples(index=False)]
        return self._rows_to_elements(rows, sheet_name="csv")

    # ---- Shared helper ----

    def _rows_to_elements(
        self, rows: list[tuple], sheet_name: str = ""
    ) -> list[DocumentElement]:
        """Convert raw rows into Markdown TABLE elements, splitting large tables."""
        if not rows:
            return []

        # Clean: drop fully-empty rows
        rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
        if not rows:
            return []

        header = rows[0]
        data_rows = rows[1:]
        elements: list[DocumentElement] = []

        # Sheet heading
        if sheet_name:
            elements.append(self._build_element(
                ElementType.HEADING, f"Sheet: {sheet_name}", level=2
            ))

        # Split into segments of _MAX_ROWS_PER_ELEMENT
        for i in range(0, max(len(data_rows), 1), _MAX_ROWS_PER_ELEMENT):
            segment = data_rows[i : i + _MAX_ROWS_PER_ELEMENT]
            md = self._to_markdown_table(header, segment)
            elements.append(self._build_element(
                ElementType.TABLE, md,
                sheet=sheet_name,
                row_start=i + 1,
                row_end=i + len(segment),
            ))

        return elements

    @staticmethod
    def _to_markdown_table(header: tuple, rows: list[tuple]) -> str:
        def _cell(v):
            s = str(v) if v is not None else ""
            return s.replace("|", "\\|").replace("\n", " ")

        lines = ["| " + " | ".join(_cell(c) for c in header) + " |"]
        lines.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row in rows:
            # Pad/trim to match header length
            padded = list(row) + [None] * (len(header) - len(row))
            lines.append("| " + " | ".join(_cell(c) for c in padded[: len(header)]) + " |")
        return "\n".join(lines)
